from datetime import date

from openpyxl import Workbook
from openpyxl import load_workbook

FILE_PATH = f'\\\\miacshare\\mailzdrav$\\userzdravnew.xlsx'

def write_excel(email: str, username: str, userpassword: str, discript: str) -> None:
    """ создаём эксельку """
    #os.chdir('C:\\example')
    wb = Workbook()
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    data = [email, username, userpassword, discript, str(date.today())]
    ws.append(data)
    wb.save(FILE_PATH)
